# automation for processing spreadsheets

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')

sheet = wb['Sheet1']

# cell = sheet['a1']

# cell = sheet.cell(1, 1)

# print(cell.value)

# print(sheet.max_row)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    # print(cell.value)
    correctedPrice = cell.value * .9
    correctedCell = sheet.cell(row, 4)
    correctedCell.value = correctedPrice

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4
                   )

chart = BarChart()
chart.add_data(values)

sheet.add_chart(chart, 'E2')

wb.save('transactions2.xlsx')


